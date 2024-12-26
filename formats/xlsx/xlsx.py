# import os
# import zipfile
# import math 
# from xml.etree import ElementTree as ET
# from reportlab.lib import colors
# from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
# from reportlab.lib.pagesizes import letter, landscape
# from reportlab.lib.units import cm
# from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
# from reportlab.lib.enums import TA_CENTER

# NAMESPACE = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

# def extract_xlsx_xml(input_path: str):
#     """Extracts and processes XML files from an .xlsx file."""
#     xml_files = {}
#     with zipfile.ZipFile(input_path, 'r') as xlsx:
#         for file_name in xlsx.namelist():
#             if file_name.endswith('.xml'):
#                 with xlsx.open(file_name) as xml_file:
#                     tree = ET.parse(xml_file)
#                     xml_files[file_name] = tree
#     return xml_files

# def parse_sheet(sheet_tree):
#     """
#     Procesa el XML de una hoja de cálculo para extraer los valores de las celdas
#     y asegura que todas las filas tengan el mismo número de columnas.
#     También maneja celdas combinadas, incluyendo las combinaciones verticales.
#     """
#     root = sheet_tree.getroot()
#     rows = []
#     merged_cells = []

#     # Extrae celdas combinadas (si existen)
#     merge_cells_element = root.find('ns:mergeCells', NAMESPACE)
#     if merge_cells_element is not None:
#         for merge_cell in merge_cells_element.findall('ns:mergeCell', NAMESPACE):
#             merged_cells.append(merge_cell.attrib['ref'])

#     max_columns = 0
#     row_data_map = {}  # Mapa para manejar celdas por fila

#     for row in root.findall('ns:sheetData/ns:row', NAMESPACE):
#         row_index = int(row.attrib['r'])
#         row_data = {}
#         for cell in row.findall('ns:c', NAMESPACE):
#             cell_reference = cell.attrib['r']
#             cell_value = cell.find('ns:v', NAMESPACE)
#             row_data[cell_reference] = cell_value.text if cell_value is not None else ""

#         # Maneja las celdas combinadas
#         for merged_cell in merged_cells:
#             start_cell, end_cell = merged_cell.split(':')
#             start_col, start_row = split_cell_reference(start_cell)
#             end_col, end_row = split_cell_reference(end_cell)

#             if row_index in range(start_row, end_row + 1):
#                 merged_value = row_data.get(start_cell, "")
#                 for col in range(start_col, end_col + 1):
#                     for r in range(start_row, end_row + 1):
#                         cell_ref = f"{convert_col_index_to_letter(col)}{r}"
#                         if r == start_row and col == start_col:
#                             continue  # No sobrescribas la celda de inicio
#                         row_data[cell_ref] = merged_value

#         row_data_map[row_index] = row_data

#     # Ordenar las filas y convertirlas en listas
#     for row_index in sorted(row_data_map.keys()):
#         row_data = row_data_map[row_index]
#         max_columns = max(max_columns, len(row_data))
#         ordered_row = []
#         for col in range(1, max_columns + 1):
#             cell_ref = f"{convert_col_index_to_letter(col)}{row_index}"
#             ordered_row.append(row_data.get(cell_ref, ""))
#         rows.append(ordered_row)

#     # Asegura que todas las filas tengan el mismo número de columnas
#     for row in rows:
#         while len(row) < max_columns:
#             row.append("")

#     return rows

# def split_cell_reference(cell_ref):
#     col = ''.join(filter(str.isalpha, cell_ref))
#     row = ''.join(filter(str.isdigit, cell_ref))
#     col_index = sum((ord(char) - 64) * (26 ** i) for i, char in enumerate(reversed(col)))
#     return col_index, int(row)

# def convert_col_index_to_letter(index):
#     letter = ""
#     while index > 0:
#         index, remainder = divmod(index - 1, 26)
#         letter = chr(65 + remainder) + letter
#     return letter

# def get_shared_strings(xml_files):
#     shared_strings = []
#     shared_strings_file = xml_files.get('xl/sharedStrings.xml')
#     if shared_strings_file:
#         root = shared_strings_file.getroot()
#         for si in root.findall('ns:si', NAMESPACE):
#             t_element = si.find('ns:t', NAMESPACE)
#             shared_strings.append(t_element.text if t_element is not None else "")
#     return shared_strings

# def calculate_column_widths(data, page_width, margin=2 * cm):
#     if not data or not any(data):
#         return []
        
#     # Calcular longitudes máximas como antes
#     max_lengths = [max(len(str(cell)) for cell in col) for col in zip(*data)]
#     total_length = sum(max_lengths) or 1
    
#     # Calcular el ancho disponible
#     usable_width = page_width - 2 * margin
    
#     # Establecer un ancho mínimo por columna (por ejemplo, 1 cm)
#     min_column_width = 1 * cm
    
#     # Calcular anchos proporcionales con mínimo
#     column_widths = []
#     for length in max_lengths:
#         width = (length / total_length) * usable_width
#         column_widths.append(max(width, min_column_width))
        
#     # Ajustar si el total excede el ancho disponible
#     if sum(column_widths) > usable_width:
#         scale_factor = usable_width / sum(column_widths)
#         column_widths = [width * scale_factor for width in column_widths]
        
#     return column_widths

# def process_cell(value, centered_style, column_width, base_font_size=10, min_font_size=6):
#     """
#     Procesa una celda adaptando su tamaño al contenido y asegurando que el texto
#     se ajuste correctamente dentro de ella.
#     """
#     if value is None or value == "":
#         value = " "
    
#     # Limpiamos y normalizamos el valor
#     value = str(value).strip()
#     value = value.replace('\x00', '')
    
#     # Establecemos un ancho mínimo razonable
#     min_width = 1 * cm
#     column_width = max(column_width, min_width)
    
#     try:
#         # Calculamos el tamaño de fuente óptimo y las dimensiones necesarias
#         optimal_font_size = base_font_size
#         width, height = calculate_text_dimensions(value, optimal_font_size, column_width)
        
#         # Si el texto es muy largo, reducimos el tamaño de la fuente gradualmente
#         while height > 4 * optimal_font_size and optimal_font_size > min_font_size:
#             optimal_font_size -= 1
#             width, height = calculate_text_dimensions(value, optimal_font_size, column_width)
        
#         # Creamos un estilo personalizado con las dimensiones calculadas
#         custom_style = ParagraphStyle(
#             'custom',
#             parent=centered_style,
#             fontSize=optimal_font_size,
#             leading=optimal_font_size * 1.2,  # Espaciado entre líneas
#             wordWrap='CJK',
#             alignment=TA_CENTER,
#             allowWidows=1,
#             allowOrphans=1,
#             spaceBefore=3,  # Añadimos un pequeño espacio antes
#             spaceAfter=3    # y después del texto
#         )
        
#         # Creamos el párrafo con el texto formateado
#         return Paragraph(value, custom_style), height
        
#     except Exception as e:
#         # En caso de error, retornamos una celda simple con altura mínima
#         return Paragraph(" ", centered_style), base_font_size * 1.2
    
# def calculate_row_heights(data, column_widths, centered_style, base_font_size=10, min_font_size=6, max_lines=3):
#     """
#     Calcula la altura necesaria para cada fila basándose en el contenido
#     de todas sus celdas.

#     Parámetros:
#     - data: Lista de listas con los datos de las filas.
#     - column_widths: Lista con los anchos de las columnas.
#     - centered_style: Estilo base para el texto.
#     - base_font_size: Tamaño base de la fuente.
#     - min_font_size: Tamaño mínimo de la fuente.
#     - max_lines: Máximo número de líneas permitido por celda.

#     Retorna:
#     - Lista con las alturas calculadas para cada fila.
#     """
#     row_heights = []
#     for row in data:
#         max_row_height = 0
#         for cell, width in zip(row, column_widths):
#             # Procesa cada celda para obtener su altura
#             _, cell_height = process_cell(cell, centered_style, width, base_font_size, min_font_size)
#             max_row_height = max(max_row_height, cell_height)
#         # Agrega un pequeño padding a la altura de la fila
#         row_heights.append(max_row_height + 6)
#     return row_heights

# def adjust_font_size_and_height(value, column_width, base_font_size=10, min_font_size=6, max_lines=3):
#     """
#     Ajusta dinámicamente el tamaño de la fuente y calcula el alto necesario para la celda.
#     """
#     if column_width <= 0:
#         return min_font_size, max_lines * min_font_size * 1.2

#     current_font_size = base_font_size
#     text = str(value)
#     while current_font_size >= min_font_size:
#         chars_per_line = max(1, int(column_width / (current_font_size * 0.6)))
#         required_lines = (len(text) // chars_per_line) + 1

#         if required_lines <= max_lines:
#             return current_font_size, required_lines * current_font_size * 1.2

#         current_font_size -= 1

#     return min_font_size, max_lines * min_font_size * 1.2

# def calculate_text_dimensions(text, font_size, column_width):
#     """
#     Calcula las dimensiones aproximadas que ocupará un texto dado un tamaño de fuente
#     y ancho de columna. Esta función ayuda a determinar cuánto espacio necesitará realmente
#     el texto en la celda.
#     """
#     if not text or text.isspace():
#         return column_width, font_size

#     # Calculamos caracteres aproximados por línea
#     chars_per_line = max(1, int(column_width / (font_size * 0.6)))
    
#     # Calculamos cuántas líneas necesitaríamos
#     text_length = len(str(text))
#     num_lines = max(1, math.ceil(text_length / chars_per_line))
    
#     # Altura total necesaria (considerando espaciado entre líneas)
#     height = num_lines * (font_size * 1.2)  # 1.2 es el factor de espaciado entre líneas
    
#     return column_width, height

# def extract_cell_styles(xml_files):
#     """
#     Extrae los estilos de las celdas desde el archivo styles.xml.
#     """
#     styles = {}
#     styles_file = xml_files.get('xl/styles.xml')
#     if styles_file is None:
#         return styles  # No hay estilos disponibles

#     root = styles_file.getroot()
#     ns = {'a': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

#     # Extraer fuentes
#     fonts = root.find('a:fonts', ns)
#     if fonts is not None:
#         for idx, font in enumerate(fonts.findall('a:font', ns)):
#             font_size = font.find('a:sz', ns)
#             color = font.find('a:color', ns)
#             bold = font.find('a:b', ns) is not None
#             italic = font.find('a:i', ns) is not None

#             styles[idx] = {
#                 'font_size': float(font_size.attrib.get('val', 10)) if font_size is not None else 10,
#                 'color': f"#{color.attrib.get('rgb', '000000')}" if color is not None else '#000000',
#                 'bold': bold,
#                 'italic': italic
#             }

#     return styles

# def convert_xlsx_to_pdf(input_path: str, output_path: str) -> None:
#     xml_files = extract_xlsx_xml(input_path)
#     shared_strings = get_shared_strings(xml_files)
#     sheets_data = {}

#     for file_name, tree in xml_files.items():
#         if file_name.startswith('xl/worksheets/sheet'):
#             sheet_name = file_name.split('/')[-1].replace('.xml', '')
#             sheets_data[sheet_name] = parse_sheet(tree)

#     styles = getSampleStyleSheet()
#     centered_style = ParagraphStyle(
#         'centered',
#         parent=styles['Normal'],
#         alignment=TA_CENTER,
#         wordWrap='CJK',
#         fontSize=10,
#         leading=12
#     )

#     pdf = SimpleDocTemplate(
#         output_path,
#         pagesize=landscape(letter),
#         leftMargin=2 * cm,
#         rightMargin=2 * cm,
#         topMargin=2 * cm,
#         bottomMargin=2 * cm
#     )

#     page_width, page_height = landscape(letter)
#     elements = []

#     for sheet_name, rows in sheets_data.items():
#         if not rows:
#             continue

#         data = []
#         column_widths = calculate_column_widths(rows, page_width)
#         processed_cells = []

#         for row in rows:
#             processed_row = []
#             for col_idx, cell in enumerate(row):
#                 cell_content = (
#                     shared_strings[int(cell)] if cell.isdigit() and int(cell) < len(shared_strings) else cell
#                 )
#                 processed_cell, _ = process_cell(
#                     cell_content,
#                     centered_style,
#                     column_widths[col_idx]
#                 )
#                 processed_row.append(processed_cell)
#             processed_cells.append(processed_row)

#         # Calcula las alturas de las filas, pasando `centered_style` como argumento
#         row_heights = calculate_row_heights(rows, column_widths, centered_style)

#         table = Table(processed_cells, colWidths=column_widths, rowHeights=row_heights)
#         style = TableStyle([
#             ('GRID', (0, 0), (-1, -1), 1, colors.black),
#             ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
#             ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
#             ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
#             ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
#             ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
#             ('TOPPADDING', (0, 0), (-1, -1), 3),
#             ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
#             ('LEFTPADDING', (0, 0), (-1, -1), 3),
#             ('RIGHTPADDING', (0, 0), (-1, -1), 3),
#         ])
#         table.setStyle(style)

#         elements.append(Paragraph(f"Hoja: {sheet_name}", styles['Heading2']))
#         elements.append(table)
#         elements.append(Spacer(1, 12))

#     try:
#         pdf.build(elements)
#     except Exception as e:
#         raise ValueError(f"CONVERSION_ERROR: {e}")
    
import io
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl.utils import get_column_letter
from reportlab.lib.units import cm

def get_column_widths(sheet, max_width):
    """
    Calcula los anchos de las columnas de una hoja de Excel y los escala para ajustarlos a un ancho máximo.

    Parámetros:
    - sheet: Hoja de cálculo activa (openpyxl Worksheet).
    - max_width: Ancho máximo permitido para todas las columnas.

    Retorna:
    - list[float]: Lista con los anchos escalados de cada columna.
    """
    column_widths = []
    for col in sheet.columns:
        width = sheet.column_dimensions[col[0].column_letter].width
        if width is None:
            width = 10
        column_widths.append(width * 8)

    total_width = sum(column_widths)
    if total_width > max_width:
        scale = max_width / total_width
        column_widths = [width * scale for width in column_widths]

    return column_widths

def get_row_heights(sheet, max_height=50):
    """
    Calcula las alturas de las filas de una hoja de Excel, ajustándolas a un máximo.

    Parámetros:
    - sheet: Hoja de cálculo activa (openpyxl Worksheet).
    - max_height: Altura máxima permitida para las filas.

    Retorna:
    - list[float]: Lista con las alturas ajustadas de las filas.
    """
    row_heights = []
    for row in sheet.iter_rows():
        height = sheet.row_dimensions[row[0].row].height
        if height is None:
            height = 15
        if height > max_height: 
            height = max_height
        row_heights.append(height)

    return row_heights

def process_cell(cell, centered_style, max_characters=50):
    """
    Procesa el contenido de una celda y lo ajusta a un formato apropiado para el PDF.

    Parámetros:
    - cell: Celda a procesar (openpyxl Cell).
    - centered_style: Estilo centrado para aplicar a textos largos (ReportLab ParagraphStyle).
    - max_characters: Número máximo de caracteres permitidos en una celda.

    Retorna:
    - ReportLab Image o Paragraph: Contenido procesado como una imagen o un párrafo.
    - str: Cadena vacía si la celda está vacía.
    """
    if cell.data_type == 'i':  # Image
        img = Image(io.BytesIO(cell.value))
        img.drawHeight = 50  
        img.drawWidth = 50 
        return img
    elif cell.value is not None:
        cell_value = str(cell.value)
        if len(cell_value) > max_characters: 
            cell_value = cell_value[:max_characters] + '...'  # Truncar contenido largo
        return Paragraph(cell_value, centered_style)
    return ''

def convert_xlsx_to_pdf(input_path: str, output_path: str) -> None:
    """
    Convierte un archivo XLSX (Excel) al formato PDF.

    Parámetros:
    - input_path (str): Ruta del archivo XLSX de entrada.
    - output_path (str): Ruta donde se generará el archivo PDF.

    Retorna:
    - None.

    Lanza:
    - ValueError: Si ocurre un error durante la conversión.
    """
    wb = load_workbook(input_path)
    sheet = wb.active
    sheet_width = 0
    for column in sheet.columns:
        col_letter = get_column_letter(column[0].column)
        col_width = sheet.column_dimensions[col_letter].width
        if col_width is None:
            col_width = 8.43  # Ancho predeterminado en Excel (8.43 caracteres)
        sheet_width += col_width
    
    sheet_height = 0
    for row in sheet.rows:
        row_height = sheet.row_dimensions[row[0].row].height
        if row_height is None:
            row_height = 15  
        sheet_height += row_height
        
    # Convertir las dimensiones a puntos (1 punto = 1/72 pulgadas)
    page_width = sheet_width * 0.5  # Aproximación de unidades de Excel a cm
    page_height = sheet_height * 0.05  # Aproximación de unidades de Excel a cm
    margin_left = 0.1 * cm
    margin_right = 0.1 * cm
    margin_top = 0.2 * cm
    margin_bottom = 0.2 * cm
    pdf = SimpleDocTemplate(
        output_path,
        pagesize=(page_width * cm, page_height * cm),
        leftMargin=margin_left,
        rightMargin=margin_right,
        topMargin=margin_top,
        bottomMargin=margin_bottom
    )
    
    data = []
    styles = getSampleStyleSheet()
    centered_style = ParagraphStyle('centered', parent=styles['Normal'], alignment=TA_CENTER, wordWrap='CJK')
    
    for row in sheet.iter_rows():
        processed_row = []
        for cell in row:
            processed_cell = process_cell(cell, centered_style)
            processed_row.append(processed_cell)
        data.append(processed_row)
    column_widths = get_column_widths(sheet, page_width * cm - (margin_left + margin_right))
    table = Table(data, colWidths=column_widths, repeatRows=1)
    
    style = TableStyle([
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ])
    
    # Combinar celdas
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        style.add('SPAN', (min_col - 1, min_row - 1), (max_col - 1, max_row - 1))
    table.setStyle(style)
    elements = [table]
    try:
        pdf.build(elements)
    except Exception as e:
        raise ValueError(f"CONVERTION_ERROR: {e}")